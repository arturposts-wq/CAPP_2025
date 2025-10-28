import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QComboBox, QPushButton, QTableWidget, QTableWidgetItem, 
                             QFileDialog, QMessageBox, QDialog, QFormLayout, QTabWidget, 
                             QInputDialog, QLineEdit, QDoubleSpinBox, QSpacerItem, QSizePolicy, 
                             QGroupBox, QScrollArea, QFrame)
from PyQt5.QtCore import Qt
import sqlite3
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from datetime import datetime
import openpyxl
import traceback


class CAPPDatabase:
    def __init__(self, db_name="capp.db"):
        try:
            self.conn = sqlite3.connect(db_name)
            self.cursor = self.conn.cursor()
            self.create_tables()
            print("База данных успешно инициализирована.")
        except sqlite3.Error as e:
            print(f"Ошибка базы данных: {e}")
            raise

    def create_tables(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS models (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS parts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                name TEXT NOT NULL,
                code TEXT,
                quantity INTEGER,
                FOREIGN KEY (model_id) REFERENCES models(id)
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                number TEXT,
                code TEXT,
                name TEXT NOT NULL,
                description TEXT,
                equipment TEXT,
                document TEXT,
                prep_time REAL DEFAULT 0.0,
                unit_time REAL DEFAULT 0.0,
                FOREIGN KEY (model_id) REFERENCES models(id)
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS workshop (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                workshop_name TEXT NOT NULL,
                section TEXT,
                rm TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                article TEXT,
                note TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS document_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                organization TEXT NOT NULL,
                product_code TEXT,
                document_code TEXT,
                developed_by TEXT,
                checked_by TEXT,
                FOREIGN KEY (model_id) REFERENCES models(id)
            )
        ''')
        self.conn.commit()

    def insert_model(self, name):
        try:
            self.cursor.execute("INSERT OR IGNORE INTO models (name) VALUES (?)", (name,))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки модели: {e}")
            return None

    def insert_part(self, model_id, name, code, quantity):
        try:
            self.cursor.execute("INSERT INTO parts (model_id, name, code, quantity) VALUES (?, ?, ?, ?)",
                               (model_id, name, code, quantity))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки детали: {e}")
            return None

    def insert_operation(self, model_id, number, code, name, description, equipment="", prep_time=0.0, unit_time=0.0):
        try:
            self.cursor.execute("""
                INSERT INTO operations (model_id, number, code, name, description, equipment, prep_time, unit_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (model_id, number, code, name, description, equipment, prep_time, unit_time))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки операции: {e}")
            return None

    def insert_workshop(self, workshop_name, section, rm):
        try:
            self.cursor.execute("INSERT INTO workshop (workshop_name, section, rm) VALUES (?, ?, ?)",
                               (workshop_name, section, rm))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки данных расцеховки: {e}")
            return None

    def insert_equipment(self, name, article, note):
        try:
            self.cursor.execute("INSERT INTO equipment (name, article, note) VALUES (?, ?, ?)",
                               (name, article, note))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки оборудования: {e}")
            return None

    def insert_document_details(self, model_id, organization, product_code, document_code, developed_by, checked_by):
        try:
            self.cursor.execute("INSERT INTO document_details (model_id, organization, product_code, document_code, developed_by, checked_by) VALUES (?, ?, ?, ?, ?, ?)",
                               (model_id, organization, product_code, document_code, developed_by, checked_by))
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.Error as e:
            print(f"Ошибка вставки реквизитов документа: {e}")
            return None

    def update_model(self, id, name):
        try:
            self.cursor.execute("UPDATE models SET name = ? WHERE id = ?", (name, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления модели: {e}")
            return False

    def update_part(self, id, name, code, quantity):
        try:
            self.cursor.execute("UPDATE parts SET name = ?, code = ?, quantity = ? WHERE id = ?", (name, code, quantity, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления детали: {e}")
            return False

    def update_operation(self, id, number, code, name, description, equipment="", prep_time=0.0, unit_time=0.0):
        try:
            self.cursor.execute("""
                UPDATE operations SET number = ?, code = ?, name = ?, description = ?, equipment = ?, prep_time = ?, unit_time = ?
                WHERE id = ?
            """, (number, code, name, description, equipment, prep_time, unit_time, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления операции: {e}")
            return False

    def update_workshop(self, id, workshop_name, section, rm):
        try:
            self.cursor.execute("UPDATE workshop SET workshop_name = ?, section = ?, rm = ? WHERE id = ?", 
                               (workshop_name, section, rm, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления данных расцеховки: {e}")
            return False

    def update_equipment(self, id, name, article, note):
        try:
            self.cursor.execute("UPDATE equipment SET name = ?, article = ?, note = ? WHERE id = ?", 
                               (name, article, note, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления оборудования: {e}")
            return False

    def update_document_details(self, id, organization, product_code, document_code, developed_by, checked_by):
        try:
            self.cursor.execute("UPDATE document_details SET organization = ?, product_code = ?, document_code = ?, developed_by = ?, checked_by = ? WHERE id = ?",
                               (organization, product_code, document_code, developed_by, checked_by, id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка обновления реквизитов документа: {e}")
            return False

    def delete_model(self, id):
        try:
            self.cursor.execute("DELETE FROM parts WHERE model_id = ?", (id,))
            self.cursor.execute("DELETE FROM operations WHERE model_id = ?", (id,))
            self.cursor.execute("DELETE FROM document_details WHERE model_id = ?", (id,))
            self.cursor.execute("DELETE FROM models WHERE id = ?", (id,))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print(f"Ошибка удаления модели: {e}")
            return False

    def delete_part(self, id):
        try:
            self.cursor.execute("DELETE FROM parts WHERE id = ?", (id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка удаления детали: {e}")
            return False

    def delete_operation(self, id):
        try:
            self.cursor.execute("DELETE FROM operations WHERE id = ?", (id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка удаления операции: {e}")
            return False

    def delete_workshop(self, id):
        try:
            self.cursor.execute("DELETE FROM workshop WHERE id = ?", (id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка удаления данных расцеховки: {e}")
            return False

    def delete_equipment(self, id):
        try:
            self.cursor.execute("DELETE FROM equipment WHERE id = ?", (id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка удаления оборудования: {e}")
            return False

    def delete_document_details(self, id):
        try:
            self.cursor.execute("DELETE FROM document_details WHERE id = ?", (id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Ошибка удаления реквизитов документа: {e}")
            return False

    def get_models(self):
        try:
            self.cursor.execute("SELECT id, name FROM models")
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения моделей: {e}")
            return []

    def get_model_id(self, name):
        try:
            self.cursor.execute("SELECT id FROM models WHERE name = ?", (name,))
            result = self.cursor.fetchone()
            return result[0] if result else None
        except sqlite3.Error as e:
            print(f"Ошибка получения ID модели: {e}")
            return None

    def get_parts(self, model_id):
        try:
            self.cursor.execute("SELECT id, name, code, quantity FROM parts WHERE model_id = ?", (model_id,))
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения деталей: {e}")
            return []

    def get_operations(self, model_id=None):
        try:
            if model_id is None:
                self.cursor.execute("SELECT id, number, code, name, description, equipment, document, prep_time, unit_time FROM operations WHERE model_id IS NULL")
                return self.cursor.fetchall()
            else:
                self.cursor.execute("SELECT id, number, code, name, description, equipment, document, prep_time, unit_time FROM operations WHERE model_id = ?", (model_id,))
                return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения операций: {e}")
            return []

    def get_workshop(self):
        try:
            self.cursor.execute("SELECT id, workshop_name, section, rm FROM workshop")
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения данных расцеховки: {e}")
            return []

    def get_equipment(self):
        try:
            self.cursor.execute("SELECT id, name, article, note FROM equipment")
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения оборудования: {e}")
            return []

    def get_document_details(self, model_id):
        try:
            self.cursor.execute("SELECT id, organization, product_code, document_code, developed_by, checked_by FROM document_details WHERE model_id = ?", (model_id,))
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Ошибка получения реквизитов документа: {e}")
            return []

    def import_from_excel(self, file_path, current_model=None):
        try:
            wb = openpyxl.load_workbook(file_path)
            imported = False
            if 'Лист1' in wb.sheetnames:
                ws = wb['Лист1']
                headers = [cell.value for cell in ws[1] if cell.value]
                required = ['№', 'Номенклатура', 'Количество']
                if all(col in headers for col in required):
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        code, name, quantity = [row[headers.index(col)] for col in required]
                        if name and quantity is not None:
                            if current_model:
                                model_id = self.get_model_id(current_model)
                                if model_id:
                                    self.insert_part(model_id, name, code, quantity)
                                    imported = true
                else:
                    print("Ошибка: В листе 'Лист1' отсутствуют колонки: №, Номенклатура, Количество")
            else:
                print("Ошибка: Лист 'Лист1' не найден")
            return imported
        except Exception as e:
            print(f"Ошибка импорта из Excel: {e}")
            return False

    def close(self):
        self.conn.close()


class OperationDialog(QDialog):
    def __init__(self, parent=None, is_edit_db=False, db=None):
        super().__init__(parent)
        self.is_edit_db = is_edit_db
        self.db = db
        self.setWindowTitle("Добавить операцию")
        self.setGeometry(200, 200, 450, 250 if is_edit_db else 350)
        layout = QFormLayout()

        if not is_edit_db:
            self.number_input = QLineEdit(self)
            layout.addRow("Номер:", self.number_input)

        self.code_combo = QComboBox(self)
        self.code_combo.setEditable(True)
        self.code_combo.setStyleSheet("font-size: 14px; padding: 5px;")
        layout.addRow("Код:", self.code_combo)

        self.name_combo = QComboBox(self)
        self.name_combo.setEditable(True)
        self.name_combo.setStyleSheet("font-size: 14px; padding: 5px;")
        layout.addRow("Наименование:", self.name_combo)

        if not is_edit_db:
            self.description_input = QLineEdit(self)
            layout.addRow("Описание:", self.description_input)

            self.equipment_combo = QComboBox(self)
            self.equipment_combo.addItem("")
            if db:
                for _, name, _, _ in db.get_equipment():
                    self.equipment_combo.addItem(name)
            layout.addRow("Оборудование:", self.equipment_combo)

            self.prep_time_input = QDoubleSpinBox(self)
            self.prep_time_input.setRange(0.0, 999.9)
            self.prep_time_input.setDecimals(2)
            self.prep_time_input.setSuffix(" ч")
            layout.addRow("Время подготовки:", self.prep_time_input)

            self.unit_time_input = QDoubleSpinBox(self)
            self.unit_time_input.setRange(0.0, 999.9)
            self.unit_time_input.setDecimals(2)
            self.unit_time_input.setSuffix(" мин")
            layout.addRow("Штучное время:", self.unit_time_input)

        buttons = QHBoxLayout()
        ok_button = QPushButton("ОК", self)
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена", self)
        cancel_button.clicked.connect(self.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addLayout(buttons)
        self.setLayout(main_layout)

        if db:
            self.populate_combos()
            self.code_combo.currentTextChanged.connect(self.sync_name)
            self.name_combo.currentTextChanged.connect(self.sync_code)

    def populate_combos(self):
        operations = self.db.get_operations()
        self.code_combo.clear()
        self.name_combo.clear()
        self.operation_map = {}
        for op in operations:
            code = op[2] or ""
            name = op[3] or ""
            self.operation_map[code] = name
            self.operation_map[name] = code
            if code and self.code_combo.findText(code) == -1:
                self.code_combo.addItem(code)
            if name and self.name_combo.findText(name) == -1:
                self.name_combo.addItem(name)

    def sync_name(self, code):
        if code in self.operation_map:
            name = self.operation_map[code]
            if name and self.name_combo.currentText() != name:
                self.name_combo.setCurrentText(name)

    def sync_code(self, name):
        if name in self.operation_map:
            code = self.operation_map[name]
            if code and self.code_combo.currentText() != code:
                self.code_combo.setCurrentText(code)

    def get_values(self):
        if self.is_edit_db:
            return (self.code_combo.currentText(), self.name_combo.currentText())
        else:
            return (
                self.number_input.text(),
                self.code_combo.currentText(),
                self.name_combo.currentText(),
                self.description_input.text(),
                self.equipment_combo.currentText(),
                self.prep_time_input.value(),
                self.unit_time_input.value()
            )


class WorkshopDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить расцеховку")
        self.setGeometry(200, 200, 300, 150)
        layout = QFormLayout()

        self.workshop_input = QLineEdit(self)
        layout.addRow("Цех:", self.workshop_input)

        self.section_input = QLineEdit(self)
        layout.addRow("Участок:", self.section_input)

        self.rm_input = QLineEdit(self)
        layout.addRow("РМ:", self.rm_input)

        buttons = QHBoxLayout()
        ok_button = QPushButton("ОК", self)
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена", self)
        cancel_button.clicked.connect(self.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addLayout(buttons)
        self.setLayout(main_layout)

    def get_values(self):
        return self.workshop_input.text(), self.section_input.text(), self.rm_input.text()


class EquipmentDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить оборудование")
        self.setGeometry(200, 200, 300, 150)
        layout = QFormLayout()

        self.name_input = QLineEdit(self)
        layout.addRow("Наименование:", self.name_input)

        self.article_input = QLineEdit(self)
        layout.addRow("Артикул:", self.article_input)

        self.note_input = QLineEdit(self)
        layout.addRow("Примечание:", self.note_input)

        buttons = QHBoxLayout()
        ok_button = QPushButton("ОК", self)
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена", self)
        cancel_button.clicked.connect(self.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addLayout(buttons)
        self.setLayout(main_layout)

    def get_values(self):
        return self.name_input.text(), self.article_input.text(), self.note_input.text()


class DocumentDetailsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить реквизиты документа")
        self.setGeometry(200, 200, 400, 200)
        layout = QFormLayout()

        self.organization_input = QLineEdit(self)
        layout.addRow("Организация:", self.organization_input)

        self.product_code_input = QLineEdit(self)
        layout.addRow("Обозначение изделия:", self.product_code_input)

        self.document_code_input = QLineEdit(self)
        layout.addRow("Обозначение документа:", self.document_code_input)

        self.developed_by_input = QLineEdit(self)
        layout.addRow("Разработал:", self.developed_by_input)

        self.checked_by_input = QLineEdit(self)
        layout.addRow("Проверил:", self.checked_by_input)

        buttons = QHBoxLayout()
        ok_button = QPushButton("ОК", self)
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена", self)
        cancel_button.clicked.connect(self.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(layout)
        main_layout.addLayout(buttons)
        self.setLayout(main_layout)

    def get_values(self):
        return (self.organization_input.text(), self.product_code_input.text(), 
                self.document_code_input.text(), self.developed_by_input.text(), 
                self.checked_by_input.text())


class EditDBDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Редактировать БД")
        self.setGeometry(100, 100, 800, 600)
        layout = QVBoxLayout()

        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)

        self.operations_tab = QWidget()
        self.tab_widget.addTab(self.operations_tab, "Операции")
        self.setup_operations_tab()

        self.workshop_tab = QWidget()
        self.tab_widget.addTab(self.workshop_tab, "Расцеховка")
        self.setup_workshop_tab()

        self.equipment_tab = QWidget()
        self.tab_widget.addTab(self.equipment_tab, "Оборудование")
        self.setup_equipment_tab()

        buttons = QHBoxLayout()
        close_button = QPushButton("Закрыть")
        close_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #FF9800; color: white; border-radius: 5px;")
        close_button.clicked.connect(self.accept)
        buttons.addWidget(close_button)
        layout.addLayout(buttons)
        self.setLayout(layout)

    def setup_operations_tab(self):
        layout = QVBoxLayout()
        self.operations_table = QTableWidget()
        self.operations_table.setColumnCount(2)
        self.operations_table.setHorizontalHeaderLabels(['Код', 'Наименование'])
        self.operations_table.horizontalHeader().setStretchLastSection(True)
        self.operations_table.setStyleSheet("font-size: 14px; color: #333;")
        layout.addWidget(self.operations_table)
        buttons = QHBoxLayout()
        add_button = QPushButton("Добавить")
        add_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #4CAF50; color: white; border-radius: 5px;")
        add_button.clicked.connect(self.add_operation)
        buttons.addWidget(add_button)
        edit_operation_button = QPushButton("Редактировать")
        edit_operation_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #2196F3; color: white; border-radius: 5px;")
        edit_operation_button.clicked.connect(self.edit_operation)
        buttons.addWidget(edit_operation_button)
        delete_button = QPushButton("Удалить")
        delete_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #F44336; color: white; border-radius: 5px;")
        delete_button.clicked.connect(self.delete_operation)
        buttons.addWidget(delete_button)
        layout.addLayout(buttons)
        self.operations_tab.setLayout(layout)
        self.update_operations_table()

    def setup_workshop_tab(self):
        layout = QVBoxLayout()
        self.workshop_table = QTableWidget()
        self.workshop_table.setColumnCount(3)
        self.workshop_table.setHorizontalHeaderLabels(['Цех', 'Участок', 'РМ'])
        self.workshop_table.horizontalHeader().setStretchLastSection(True)
        self.workshop_table.setStyleSheet("font-size: 14px; color: #333;")
        layout.addWidget(self.workshop_table)
        buttons = QHBoxLayout()
        add_button = QPushButton("Добавить")
        add_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #4CAF50; color: white; border-radius: 5px;")
        add_button.clicked.connect(self.add_workshop)
        buttons.addWidget(add_button)
        edit_button = QPushButton("Редактировать")
        edit_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #2196F3; color: white; border-radius: 5px;")
        edit_button.clicked.connect(self.edit_workshop)
        buttons.addWidget(edit_button)
        delete_button = QPushButton("Удалить")
        delete_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #F44336; color: white; border-radius: 5px;")
        delete_button.clicked.connect(self.delete_workshop)
        buttons.addWidget(delete_button)
        layout.addLayout(buttons)
        self.workshop_tab.setLayout(layout)
        self.update_workshop_table()

    def setup_equipment_tab(self):
        layout = QVBoxLayout()
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(3)
        self.equipment_table.setHorizontalHeaderLabels(['Наименование', 'Артикул', 'Примечание'])
        self.equipment_table.horizontalHeader().setStretchLastSection(True)
        self.equipment_table.setStyleSheet("font-size: 14px; color: #333;")
        layout.addWidget(self.equipment_table)
        buttons = QHBoxLayout()
        add_button = QPushButton("Добавить")
        add_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #4CAF50; color: white; border-radius: 5px;")
        add_button.clicked.connect(self.add_equipment)
        buttons.addWidget(add_button)
        edit_button = QPushButton("Редактировать")
        edit_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #2196F3; color: white; border-radius: 5px;")
        edit_button.clicked.connect(self.edit_equipment)
        buttons.addWidget(edit_button)
        delete_button = QPushButton("Удалить")
        delete_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #F44336; color: white; border-radius: 5px;")
        delete_button.clicked.connect(self.delete_equipment)
        buttons.addWidget(delete_button)
        layout.addLayout(buttons)
        self.equipment_tab.setLayout(layout)
        self.update_equipment_table()

    def update_operations_table(self):
        operations = self.db.get_operations()
        self.operations_table.setRowCount(len(operations))
        for row, (id, _, code, name, _, _, _, _, _) in enumerate(operations):
            self.operations_table.setItem(row, 0, QTableWidgetItem(code or ""))
            self.operations_table.setItem(row, 1, QTableWidgetItem(name))
            self.operations_table.item(row, 0).setData(Qt.UserRole, id)

    def update_workshop_table(self):
        workshops = self.db.get_workshop()
        self.workshop_table.setRowCount(len(workshops))
        for row, (id, workshop_name, section, rm) in enumerate(workshops):
            self.workshop_table.setItem(row, 0, QTableWidgetItem(workshop_name))
            self.workshop_table.setItem(row, 1, QTableWidgetItem(section or ""))
            self.workshop_table.setItem(row, 2, QTableWidgetItem(rm or ""))
            self.workshop_table.item(row, 0).setData(Qt.UserRole, id)

    def update_equipment_table(self):
        equipment = self.db.get_equipment()
        self.equipment_table.setRowCount(len(equipment))
        for row, (id, name, article, note) in enumerate(equipment):
            self.equipment_table.setItem(row, 0, QTableWidgetItem(name))
            self.equipment_table.setItem(row, 1, QTableWidgetItem(article or ""))
            self.equipment_table.setItem(row, 2, QTableWidgetItem(note or ""))
            self.equipment_table.item(row, 0).setData(Qt.UserRole, id)

    def add_operation(self):
        dialog = OperationDialog(self, is_edit_db=True, db=self.db)
        if dialog.exec_() == QDialog.Accepted:
            code, name = dialog.get_values()
            if name:
                # Добавление в справочник операций
                self.db.cursor.execute("INSERT OR IGNORE INTO operations (model_id, code, name) VALUES (NULL, ?, ?)", (code, name))
                self.db.conn.commit()
                self.update_operations_table()
                QMessageBox.information(self, "Успех", f"Операция '{name}' добавлена в справочник!")
            else:
                QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Наименование'!")

    def edit_operation(self):
        row = self.operations_table.currentRow()
        if row >= 0:
            old_code = self.operations_table.item(row, 0).text()
            old_name = self.operations_table.item(row, 1).text()
            dialog = OperationDialog(self, is_edit_db=True, db=self.db)
            dialog.code_combo.setCurrentText(old_code)
            dialog.name_combo.setCurrentText(old_name)
            if dialog.exec_() == QDialog.Accepted:
                code, name = dialog.get_values()
                if name:
                    op_id = self.operations_table.item(row, 0).data(Qt.UserRole)
                    self.db.cursor.execute("UPDATE operations SET code = ?, name = ? WHERE id = ?", (code, name, op_id))
                    self.db.conn.commit()
                    self.update_operations_table()
                    QMessageBox.information(self, "Успех", "Операция обновлена!")
                else:
                    QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Наименование'!")

    def delete_operation(self):
        row = self.operations_table.currentRow()
        if row >= 0:
            name = self.operations_table.item(row, 1).text()
            reply = QMessageBox.question(self, "Подтверждение", f"Удалить операцию '{name}' из справочника?", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                op_id = self.operations_table.item(row, 0).data(Qt.UserRole)
                self.db.delete_operation(op_id)
                self.update_operations_table()
                QMessageBox.information(self, "Успех", "Операция удалена!")

    def add_workshop(self):
        dialog = WorkshopDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            workshop_name, section, rm = dialog.get_values()
            if workshop_name:
                self.db.insert_workshop(workshop_name, section, rm)
                self.update_workshop_table()
                QMessageBox.information(self, "Успех", f"Цех '{workshop_name}' добавлен!")
            else:
                QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Цех'!")

    def edit_workshop(self):
        row = self.workshop_table.currentRow()
        if row >= 0:
            old_name = self.workshop_table.item(row, 0).text()
            dialog = WorkshopDialog(self)
            dialog.workshop_input.setText(old_name)
            dialog.section_input.setText(self.workshop_table.item(row, 1).text())
            dialog.rm_input.setText(self.workshop_table.item(row, 2).text())
            if dialog.exec_() == QDialog.Accepted:
                workshop_name, section, rm = dialog.get_values()
                if workshop_name:
                    ws_id = self.workshop_table.item(row, 0).data(Qt.UserRole)
                    self.db.update_workshop(ws_id, workshop_name, section, rm)
                    self.update_workshop_table()
                    QMessageBox.information(self, "Успех", "Данные обновлены!")
                else:
                    QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Цех'!")

    def delete_workshop(self):
        row = self.workshop_table.currentRow()
        if row >= 0:
            name = self.workshop_table.item(row, 0).text()
            reply = QMessageBox.question(self, "Подтверждение", f"Удалить '{name}'?", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                ws_id = self.workshop_table.item(row, 0).data(Qt.UserRole)
                self.db.delete_workshop(ws_id)
                self.update_workshop_table()
                QMessageBox.information(self, "Успех", "Удалено!")

    def add_equipment(self):
        dialog = EquipmentDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            name, article, note = dialog.get_values()
            if name:
                self.db.insert_equipment(name, article, note)
                self.update_equipment_table()
                QMessageBox.information(self, "Успех", f"Оборудование '{name}' добавлено!")
            else:
                QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Наименование'!")

    def edit_equipment(self):
        row = self.equipment_table.currentRow()
        if row >= 0:
            old_name = self.equipment_table.item(row, 0).text()
            dialog = EquipmentDialog(self)
            dialog.name_input.setText(old_name)
            dialog.article_input.setText(self.equipment_table.item(row, 1).text())
            dialog.note_input.setText(self.equipment_table.item(row, 2).text())
            if dialog.exec_() == QDialog.Accepted:
                name, article, note = dialog.get_values()
                if name:
                    eq_id = self.equipment_table.item(row, 0).data(Qt.UserRole)
                    self.db.update_equipment(eq_id, name, article, note)
                    self.update_equipment_table()
                    QMessageBox.information(self, "Успех", "Оборудование обновлено!")
                else:
                    QMessageBox.warning(self, "Предупреждение", "Заполните поле 'Наименование'!")

    def delete_equipment(self):
        row = self.equipment_table.currentRow()
        if row >= 0:
            name = self.equipment_table.item(row, 0).text()
            reply = QMessageBox.question(self, "Подтверждение", f"Удалить '{name}'?", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                eq_id = self.equipment_table.item(row, 0).data(Qt.UserRole)
                self.db.delete_equipment(eq_id)
                self.update_equipment_table()
                QMessageBox.information(self, "Успех", "Удалено!")


class EditTPDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Редактировать ТП")
        self.setGeometry(100, 100, 1000, 700)
        layout = QVBoxLayout()

        top_layout = QHBoxLayout()
        model_label = QLabel("Модель:")
        self.model_combo = QComboBox()
        self.model_combo.addItems([name for _, name in db.get_models()])
        top_layout.addWidget(model_label)
        top_layout.addWidget(self.model_combo)

        add_model_btn = QPushButton("Добавить")
        add_model_btn.clicked.connect(self.add_model)
        top_layout.addWidget(add_model_btn)

        layout.addLayout(top_layout)

        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)

        self.parts_tab = QWidget()
        self.tab_widget.addTab(self.parts_tab, "Спецификация")
        self.setup_parts_tab()

        self.operations_tab = QWidget()
        self.tab_widget.addTab(self.operations_tab, "Операции")
        self.setup_operations_tab()

        self.document_details_tab = QWidget()
        self.tab_widget.addTab(self.document_details_tab, "Реквизиты")
        self.setup_document_details_tab()

        self.model_combo.currentTextChanged.connect(self.load_model_data)

        buttons = QHBoxLayout()
        close_button = QPushButton("Закрыть")
        close_button.setStyleSheet("font-size: 14px; padding: 8px; background-color: #FF9800; color: white; border-radius: 5px;")
        close_button.clicked.connect(self.accept)
        buttons.addWidget(close_button)
        layout.addLayout(buttons)
        self.setLayout(layout)

        if self.model_combo.count() > 0:
            self.load_model_data(self.model_combo.currentText())

    def setup_parts_tab(self):
        layout = QVBoxLayout()
        self.parts_table = QTableWidget()
        self.parts_table.setColumnCount(4)
        self.parts_table.setHorizontalHeaderLabels(['ID', 'Номер', 'Номенклатура', 'Кол-во'])
        self.parts_table.hideColumn(0)
        layout.addWidget(self.parts_table)

        buttons = QHBoxLayout()
        add_btn = QPushButton("Добавить")
        add_btn.clicked.connect(self.add_part)
        buttons.addWidget(add_btn)
        edit_btn = QPushButton("Редактировать")
        edit_btn.clicked.connect(self.edit_part)
        buttons.addWidget(edit_btn)
        delete_btn = QPushButton("Удалить")
        delete_btn.clicked.connect(self.delete_part)
        buttons.addWidget(delete_btn)
        layout.addLayout(buttons)
        self.parts_tab.setLayout(layout)

    def setup_operations_tab(self):
        layout = QVBoxLayout()
        self.operations_table = QTableWidget()
        self.operations_table.setColumnCount(8)
        self.operations_table.setHorizontalHeaderLabels(['ID', '№', 'Код', 'Наименование', 'Описание', 'Оборудование', 'Tподг, ч', 'Tшт, мин'])
        self.operations_table.hideColumn(0)
        layout.addWidget(self.operations_table)

        buttons = QHBoxLayout()
        add_btn = QPushButton("Добавить")
        add_btn.clicked.connect(self.add_operation)
        buttons.addWidget(add_btn)
        edit_btn = QPushButton("Редактировать")
        edit_btn.clicked.connect(self.edit_operation)
        buttons.addWidget(edit_btn)
        delete_btn = QPushButton("Удалить")
        delete_btn.clicked.connect(self.delete_operation)
        buttons.addWidget(delete_btn)
        layout.addLayout(buttons)
        self.operations_tab.setLayout(layout)

    def setup_document_details_tab(self):
        layout = QVBoxLayout()
        self.document_details_table = QTableWidget()
        self.document_details_table.setColumnCount(6)
        self.document_details_table.setHorizontalHeaderLabels(['ID', 'Организация', 'Изделие', 'Документ', 'Разработал', 'Проверил'])
        self.document_details_table.hideColumn(0)
        layout.addWidget(self.document_details_table)

        buttons = QHBoxLayout()
        add_btn = QPushButton("Добавить")
        add_btn.clicked.connect(self.add_document_details)
        buttons.addWidget(add_btn)
        edit_btn = QPushButton("Редактировать")
        edit_btn.clicked.connect(self.edit_document_details)
        buttons.addWidget(edit_btn)
        delete_btn = QPushButton("Удалить")
        delete_btn.clicked.connect(self.delete_document_details)
        buttons.addWidget(delete_btn)
        layout.addLayout(buttons)
        self.document_details_tab.setLayout(layout)

    def load_model_data(self, model_name):
        model_id = self.db.get_model_id(model_name)
        if not model_id:
            return

        # Спецификация
        parts = self.db.get_parts(model_id)
        self.parts_table.setRowCount(len(parts))
        for row, (id, name, code, qty) in enumerate(parts):
            self.parts_table.setItem(row, 0, QTableWidgetItem(str(id)))
            self.parts_table.setItem(row, 1, QTableWidgetItem(code or ""))
            self.parts_table.setItem(row, 2, QTableWidgetItem(name))
            self.parts_table.setItem(row, 3, QTableWidgetItem(str(qty)))
            self.parts_table.item(row, 0).setData(Qt.UserRole, id)

        # Операции
        operations = self.db.get_operations(model_id)
        self.operations_table.setRowCount(len(operations))
        for row, (id, number, code, name, desc, equip, _, prep, unit) in enumerate(operations):
            self.operations_table.setItem(row, 0, QTableWidgetItem(str(id)))
            self.operations_table.setItem(row, 1, QTableWidgetItem(number or ""))
            self.operations_table.setItem(row, 2, QTableWidgetItem(code or ""))
            self.operations_table.setItem(row, 3, QTableWidgetItem(name))
            self.operations_table.setItem(row, 4, QTableWidgetItem(desc or ""))
            self.operations_table.setItem(row, 5, QTableWidgetItem(equip or ""))
            self.operations_table.setItem(row, 6, QTableWidgetItem(f"{prep:.2f}"))
            self.operations_table.setItem(row, 7, QTableWidgetItem(f"{unit:.2f}"))
            self.operations_table.item(row, 0).setData(Qt.UserRole, id)

        # Реквизиты
        details = self.db.get_document_details(model_id)
        self.document_details_table.setRowCount(len(details))
        for row, (id, org, prod, doc, dev, check) in enumerate(details):
            self.document_details_table.setItem(row, 0, QTableWidgetItem(str(id)))
            self.document_details_table.setItem(row, 1, QTableWidgetItem(org))
            self.document_details_table.setItem(row, 2, QTableWidgetItem(prod or ""))
            self.document_details_table.setItem(row, 3, QTableWidgetItem(doc or ""))
            self.document_details_table.setItem(row, 4, QTableWidgetItem(dev or ""))
            self.document_details_table.setItem(row, 5, QTableWidgetItem(check or ""))
            self.document_details_table.item(row, 0).setData(Qt.UserRole, id)

    def add_model(self):
        name, ok = QInputDialog.getText(self, "Добавить модель", "Название модели:")
        if ok and name:
            if self.db.insert_model(name):
                self.model_cert_combo.addItem(name)
                self.model_combo.setCurrentText(name)
                QMessageBox.information(self, "Успех", f"Модель '{name}' добавлена!")
            else:
                QMessageBox.critical(self, "Ошибка", "Модель уже существует!")

    def add_part(self):
        model_name = self.model_combo.currentText()
        model_id = self.db.get_model_id(model_name)
        if not model_id:
            return
        name, ok = QInputDialog.getText(self, "Добавить деталь", "Название:")
        if ok and name:
            code, ok = QInputDialog.getText(self, "Добавить деталь", "Код:")
            if ok:
                quantity, ok = QInputDialog.getInt(self, "Добавить деталь", "Количество:", min=1)
                if ok:
                    self.db.insert_part(model_id, name, code, quantity)
                    self.load_model_data(model_name)

    def edit_part(self):
        row = self.parts_table.currentRow()
        if row >= 0:
            part_id = self.parts_table.item(row, 0).data(Qt.UserRole)
            name = self.parts_table.item(row, 2).text()
            code = self.parts_table.item(row, 1).text()
            qty = int(self.parts_table.item(row, 3).text())
            new_name, ok = QInputDialog.getText(self, "Редактировать", "Название:", text=name)
            if ok:
                new_code, ok = QInputDialog.getText(self, "Редактировать", "Код:", text=code)
                if ok:
                    new_qty, ok = QInputDialog.getInt(self, "Редактировать", "Количество:", value=qty, min=1)
                    if ok:
                        self.db.update_part(part_id, new_name, new_code, new_qty)
                        self.load_model_data(self.model_combo.currentText())

    def delete_part(self):
        row = self.parts_table.currentRow()
        if row >= 0:
            part_id = self.parts_table.item(row, 0).data(Qt.UserRole)
            self.db.delete_part(part_id)
            self.load_model_data(self.model_combo.currentText())

    def add_operation(self):
        model_name = self.model_combo.currentText()
        model_id = self.db.get_model_id(model_name)
        if not model_id:
            return
        dialog = OperationDialog(self, is_edit_db=False, db=self.db)
        if dialog.exec_() == QDialog.Accepted:
            number, code, name, desc, equip, prep, unit = dialog.get_values()
            if name:
                self.db.insert_operation(model_id, number, code, name, desc, equip, prep, unit)
                self.load_model_data(model_name)

    def edit_operation(self):
        row = self.operations_table.currentRow()
        if row >= 0:
            op_id = self.operations_table.item(row, 0).data(Qt.UserRole)
            dialog = OperationDialog(self, is_edit_db=False, db=self.db)
            dialog.number_input.setText(self.operations_table.item(row, 1).text())
            dialog.code_combo.setCurrentText(self.operations_table.item(row, 2).text())
            dialog.name_combo.setCurrentText(self.operations_table.item(row, 3).text())
            dialog.description_input.setText(self.operations_table.item(row, 4).text())
            dialog.equipment_combo.setCurrentText(self.operations_table.item(row, 5).text())
            dialog.prep_time_input.setValue(float(self.operations_table.item(row, 6).text() or 0))
            dialog.unit_time_input.setValue(float(self.operations_table.item(row, 7).text() or 0))
            if dialog.exec_() == QDialog.Accepted:
                number, code, name, desc, equip, prep, unit = dialog.get_values()
                if name:
                    self.db.update_operation(op_id, number, code, name, desc, equip, prep, unit)
                    self.load_model_data(self.model_combo.currentText())

    def delete_operation(self):
        row = self.operations_table.currentRow()
        if row >= 0:
            op_id = self.operations_table.item(row, 0).data(Qt.UserRole)
            self.db.delete_operation(op_id)
            self.load_model_data(self.model_combo.currentText())

    def add_document_details(self):
        model_name = self.model_combo.currentText()
        model_id = self.db.get_model_id(model_name)
        if not model_id:
            return
        dialog = DocumentDetailsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            org, prod, doc, dev, check = dialog.get_values()
            if org:
                self.db.insert_document_details(model_id, org, prod, doc, dev, check)
                self.load_model_data(model_name)

    def edit_document_details(self):
        row = self.document_details_table.currentRow()
        if row >= 0:
            det_id = self.document_details_table.item(row, 0).data(Qt.UserRole)
            dialog = DocumentDetailsDialog(self)
            dialog.organization_input.setText(self.document_details_table.item(row, 1).text())
            dialog.product_code_input.setText(self.document_details_table.item(row, 2).text())
            dialog.document_code_input.setText(self.document_details_table.item(row, 3).text())
            dialog.developed_by_input.setText(self.document_details_table.item(row, 4).text())
            dialog.checked_by_input.setText(self.document_details_table.item(row, 5).text())
            if dialog.exec_() == QDialog.Accepted:
                org, prod, doc, dev, check = dialog.get_values()
                if org:
                    self.db.update_document_details(det_id, org, prod, doc, dev, check)
                    self.load_model_data(self.model_combo.currentText())

    def delete_document_details(self):
        row = self.document_details_table.currentRow()
        if row >= 0:
            det_id = self.document_details_table.item(row, 0).data(Qt.UserRole)
            self.db.delete_document_details(det_id)
            self.load_model_data(self.model_combo.currentText())


class CAPPWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = CAPPDatabase()
        self.setWindowTitle("CAPP Prototype")
        self.setGeometry(100, 100, 900, 700)
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Кнопки
        button_layout = QHBoxLayout()
        generate_btn = QPushButton("Сгенерировать техпроцесс")
        generate_btn.setStyleSheet("font-size: 14px; background-color: #808080; color: white; border-radius: 5px;")
        generate_btn.setFixedSize(200, 40)
        generate_btn.clicked.connect(self.generate_process)
        button_layout.addWidget(generate_btn)

        edit_db_btn = QPushButton("Редактировать БД")
        edit_db_btn.setStyleSheet("font-size: 14px; background-color: #808080; color: white; border-radius: 5px;")
        edit_db_btn.setFixedSize(200, 40)
        edit_db_btn.clicked.connect(self.edit_db)
        button_layout.addWidget(edit_db_btn)

        edit_tp_btn = QPushButton("Ред. ТП")
        edit_tp_btn.setStyleSheet("font-size: 14px; background-color: #808080; color: white; border-radius: 5px;")
        edit_tp_btn.setFixedSize(200, 40)
        edit_tp_btn.clicked.connect(self.edit_tp)
        button_layout.addWidget(edit_tp_btn)

        export_btn = QPushButton("Экспортировать в PDF")
        export_btn.setStyleSheet("font-size: 14px; background-color: #808080; color: white; border-radius: 5px;")
        export_btn.setFixedSize(200, 40)
        export_btn.clicked.connect(self.export_to_pdf)
        button_layout.addWidget(export_btn)

        button_layout.addSpacerItem(QSpacerItem(20, 0, QSizePolicy.Expanding, QSizePolicy.Minimum))
        layout.addLayout(button_layout)

        # Выбор модели
        input_layout = QHBoxLayout()
        model_label = QLabel("Модель:")
        model_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #333;")
        model_label.setFixedWidth(100)
        input_layout.addWidget(model_label)
        self.model_combo = QComboBox()
        self.model_combo.setStyleSheet("font-size: 14px; padding: 5px;")
        input_layout.addWidget(self.model_combo)
        self.update_model_combo()
        layout.addLayout(input_layout)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        layout.addWidget(separator)

        # Модель
        model_group = QGroupBox("Модель")
        model_group.setStyleSheet("QGroupBox { font-size: 16px; font-weight: bold; color: #2E7D32; }")
        model_layout = QVBoxLayout()
        self.model_label = QLabel("")
        self.model_label.setStyleSheet("font-size: 14px; padding: 5px; color: #333;")
        model_layout.addWidget(self.model_label)
        model_group.setLayout(model_layout)
        layout.addWidget(model_group)

        # Спецификация
        parts_group = QGroupBox("Спецификация")
        parts_group.setStyleSheet("QGroupBox { font-size: 16px; font-weight: bold; color: #2E7D32; }")
        parts_group.setCheckable(True)
        parts_group.setChecked(True)
        parts_layout = QVBoxLayout()
        self.parts_table = QTableWidget()
        self.parts_table.setColumnCount(3)
        self.parts_table.setHorizontalHeaderLabels(['Номер', 'Номенклатура', 'Количество'])
        self.parts_table.horizontalHeader().setStretchLastSection(True)
        self.parts_table.setStyleSheet("font-size: 14px; color: #333;")
        parts_scroll = QScrollArea()
        parts_scroll.setWidgetResizable(True)
        parts_scroll.setWidget(self.parts_table)
        parts_layout.addWidget(parts_scroll)
        parts_group.setLayout(parts_layout)
        layout.addWidget(parts_group)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        layout.addWidget(separator)

        # Операции
        operations_group = QGroupBox("Операции")
        operations_group.setStyleSheet("QGroupBox { font-size: 16px; font-weight: bold; color: #2E7D32; }")
        operations_group.setCheckable(True)
        operations_group.setChecked(True)
        operations_layout = QVBoxLayout()
        self.operations_table = QTableWidget()
        self.operations_table.setColumnCount(6)
        self.operations_table.setHorizontalHeaderLabels(['№', 'Код', 'Наименование', 'Оборудование', 'Tподг, ч', 'Tшт, мин'])
        self.operations_table.horizontalHeader().setStretchLastSection(True)
        self.operations_table.setStyleSheet("font-size: 14px; color: #333;")
        operations_scroll = QScrollArea()
        operations_scroll.setWidgetResizable(True)
        operations_scroll.setWidget(self.operations_table)
        operations_layout.addWidget(operations_scroll)
        operations_group.setLayout(operations_layout)
        layout.addWidget(operations_group)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        layout.addWidget(separator)

        # Расцеховка
        workshop_group = QGroupBox("Расцеховка")
        workshop_group.setStyleSheet("QGroupBox { font-size: 16px; font-weight: bold; color: #2E7D32; }")
        workshop_group.setCheckable(True)
        workshop_group.setChecked(True)
        workshop_layout = QVBoxLayout()
        self.workshop_table = QTableWidget()
        self.workshop_table.setColumnCount(3)
        self.workshop_table.setHorizontalHeaderLabels(['Цех', 'Участок', 'РМ'])
        self.workshop_table.horizontalHeader().setStretchLastSection(True)
        self.workshop_table.setStyleSheet("font-size: 14px; color: #333;")
        workshop_scroll = QScrollArea()
        workshop_scroll.setWidgetResizable(True)
        workshop_scroll.setWidget(self.workshop_table)
        workshop_layout.addWidget(workshop_scroll)
        workshop_group.setLayout(workshop_layout)
        layout.addWidget(workshop_group)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        layout.addWidget(separator)

        # Оборудование
        equipment_group = QGroupBox("Оборудование")
        equipment_group.setStyleSheet("QGroupBox { font-size: 16px; font-weight: bold; color: #2E7D32; }")
        equipment_group.setCheckable(True)
        equipment_group.setChecked(True)
        equipment_layout = QVBoxLayout()
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(3)
        self.equipment_table.setHorizontalHeaderLabels(['Наименование', 'Артикул', 'Примечание'])
        self.equipment_table.horizontalHeader().setStretchLastSection(True)
        self.equipment_table.setStyleSheet("font-size: 14px; color: #333;")
        equipment_scroll = QScrollArea()
        equipment_scroll.setWidgetResizable(True)
        equipment_scroll.setWidget(self.equipment_table)
        equipment_layout.addWidget(equipment_scroll)
        equipment_group.setLayout(equipment_layout)
        layout.addWidget(equipment_group)

        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

    def update_model_combo(self):
        self.model_combo.clear()
        self.model_combo.addItems([name for _, name in self.db.get_models()])

    def edit_db(self):
        dialog = EditDBDialog(self.db, self)
        dialog.exec_()

    def edit_tp(self):
        dialog = EditTPDialog(self.db, self)
        dialog.exec_()
        self.update_model_combo()

    def generate_process(self):
        model = self.model_combo.currentText()
        if not model:
            QMessageBox.warning(self, "Ошибка", "Выберите модель!")
            return

        model_id = self.db.get_model_id(model)
        if not model_id:
            QMessageBox.critical(self, "Ошибка", f"Модель {model} не найдена!")
            return

        parts = self.db.get_parts(model_id)
        operations = self.db.get_operations(model_id)
        workshops = self.db.get_workshop()
        equipment = self.db.get_equipment()
        document_details = self.db.get_document_details(model_id)

        self.model_label.setText(f"Модель: {model}")

        # Спецификация
        self.parts_table.setRowCount(len(parts))
        for row, (id, name, code, qty) in enumerate(parts):
            self.parts_table.setItem(row, 0, QTableWidgetItem(code or ""))
            self.parts_table.setItem(row, 1, QTableWidgetItem(name))
            self.parts_table.setItem(row, 2, QTableWidgetItem(str(qty)))

        # Операции
        self.operations_table.setRowCount(len(operations))
        for row, (id, number, code, name, _, equip, _, prep, unit) in enumerate(operations):
            self.operations_table.setItem(row, 0, QTableWidgetItem(number or ""))
            self.operations_table.setItem(row, 1, QTableWidgetItem(code or ""))
            self.operations_table.setItem(row, 2, QTableWidgetItem(name))
            self.operations_table.setItem(row, 3, QTableWidgetItem(equip or ""))
            self.operations_table.setItem(row, 4, QTableWidgetItem(f"{prep:.2f}"))
            self.operations_table.setItem(row, 5, QTableWidgetItem(f"{unit:.2f}"))

        # Расцеховка
        self.workshop_table.setRowCount(len(workshops))
        for row, (id, w, s, r) in enumerate(workshops):
            self.workshop_table.setItem(row, 0, QTableWidgetItem(w))
            self.workshop_table.setItem(row, 1, QTableWidgetItem(s or ""))
            self.workshop_table.setItem(row, 2, QTableWidgetItem(r or ""))

        # Оборудование
        self.equipment_table.setRowCount(len(equipment))
        for row, (id, name, art, note) in enumerate(equipment):
            self.equipment_table.setItem(row, 0, QTableWidgetItem(name))
            self.equipment_table.setItem(row, 1, QTableWidgetItem(art or ""))
            self.equipment_table.setItem(row, 2, QTableWidgetItem(note or ""))

        self.process_data = {
            'model': model,
            'parts': parts,
            'operations': operations,
            'workshops': workshops,
            'equipment': equipment,
            'document_details': document_details,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

    def export_to_pdf(self):
        if not hasattr(self, 'process_data'):
            QMessageBox.warning(self, "Ошибка", "Сначала сгенерируйте техпроцесс!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить PDF",
            f"Техпроцесс_{self.process_data['model']}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            "PDF Files (*.pdf)"
        )
        if not file_path:
            return

        try:
            # --- Шрифт ---
            font_dir = os.path.join(os.path.dirname(__file__), 'fonts')
            os.makedirs(font_dir, exist_ok=True)
            font_path = os.path.join(font_dir, 'DejaVuSans.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.getcwd(), 'DejaVuSans.ttf')
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('DejaVu', font_path))
                font_name = 'DejaVu'
            else:
                font_name = 'Helvetica'
                print("Шрифт DejaVu не найден, используется Helvetica")

            # --- Стили ---
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='TitleCenter', fontName=font_name, fontSize=16, alignment=TA_CENTER, spaceAfter=20))
            styles.add(ParagraphStyle(name='Header', fontName=font_name, fontSize=12, leading=14, spaceAfter=8))
            styles.add(ParagraphStyle(name='Footer', fontName=font_name, fontSize=9, alignment=TA_RIGHT))
            styles.add(ParagraphStyle(name='CellText', fontName=font_name, fontSize=9, leading=10, alignment=TA_LEFT))

            # --- Документ ---
            doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=15*mm, rightMargin=15*mm)
            story = []

            # --- Логотип ---
            logo_path = os.path.join(os.path.dirname(__file__), 'logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=50*mm, height=20*mm)
                logo.hAlign = 'CENTER'
                story.append(logo)
                story.append(Spacer(1, 5*mm))

            # --- Заголовок ---
            story.append(Paragraph("ТЕХНОЛОГИЧЕСКИЙ ПРОЦЕСС", styles['TitleCenter']))
            story.append(Paragraph(f"Модель: <b>{self.process_data['model']}</b>", styles['TitleCenter']))
            story.append(Spacer(1, 10*mm))

            # --- Реквизиты ---
            details = self.process_data.get('document_details', [])
            if details:
                data = [["Параметр", "Значение"]]
                for _, org, prod, doc, dev, check in details:
                    data += [
                        ["Организация", org or "—"],
                        ["Обозначение изделия", prod or "—"],
                        ["Обозначение документа", doc or "—"],
                        ["Разработал", dev or "—"],
                        ["Проверил", check or "—"]
                    ]
                table = Table(data, colWidths=[50*mm, 120*mm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('FONTNAME', (0,0), (-1,0), font_name),
                    ('FONTSIZE', (0,0), (-1,0), 11),
                    ('FONTNAME', (0,1), (-1,-1), font_name),
                    ('FONTSIZE', (0,1), (-1,-1), 10),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                story.append(Paragraph("Реквизиты документа", styles['Header']))
                story.append(table)
                story.append(Spacer(1, 8*mm))

            # --- Спецификация ---
            parts = self.process_data.get('parts', [])
            if parts:
                data = [["№", "Номенклатура", "Код", "Кол-во"]]
                for i, (_, name, code, qty) in enumerate(parts, 1):
                    name_para = Paragraph(name, styles['CellText']) if len(name) > 30 else name
                    data.append([str(i), name_para, code or "—", str(qty)])
                table = Table(data, colWidths=[15*mm, 100*mm, 35*mm, 20*mm], rowHeights=12*mm)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('FONTNAME', (0,0), (-1,0), font_name),
                    ('FONTSIZE', (0,0), (-1,0), 11),
                    ('FONTNAME', (0,1), (-1,-1), font_name),
                    ('FONTSIZE', (0,1), (-1,-1), 9),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                story.append(Paragraph("Спецификация", styles['Header']))
                story.append(table)
                story.append(Spacer(1, 8*mm))

            # --- Операции ---
            operations = self.process_data.get('operations', [])
            if operations:
                data = [["№", "Код", "Наименование", "Оборудование", "Tподг, ч", "Tшт, мин"]]
                for i, (_, number, code, name, _, equip, _, prep, unit) in enumerate(operations, 1):
                    name_para = Paragraph(name, styles['CellText']) if len(name) > 25 else name
                    equip_para = Paragraph(equip, styles['CellText']) if equip and len(equip) > 20 else (equip or "—")
                    data.append([number or str(i), code or "—", name_para, equip_para, f"{prep:.2f}", f"{unit:.2f}"])
                table = Table(data, colWidths=[15*mm, 25*mm, 60*mm, 50*mm, 20*mm, 20*mm], rowHeights=14*mm)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2196F3')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (3,-1), 'CENTER'),
                    ('ALIGN', (4,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('FONTNAME', (0,0), (-1,0), font_name),
                    ('FONTSIZE', (0,0), (-1,0), 11),
                    ('FONTNAME', (0,1), (-1,-1), font_name),
                    ('FONTSIZE', (0,1), (-1,-1), 9),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                story.append(Paragraph("Операции", styles['Header']))
                story.append(table)
                story.append(Spacer(1, 8*mm))

            # --- Расцеховка ---
            workshops = self.process_data.get('workshops', [])
            if workshops:
                data = [["Цех", "Участок", "РМ"]]
                for _, w, s, r in workshops:
                    data.append([w or "—", s or "—", r or "—"])
                table = Table(data, colWidths=[60*mm, 60*mm, 60*mm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FF9800')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), font_name),
                    ('FONTSIZE', (0,0), (-1,0), 11),
                    ('FONTNAME', (0,1), (-1,-1), font_name),
                    ('FONTSIZE', (0,1), (-1,-1), 10),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                story.append(Paragraph("Расцеховка", styles['Header']))
                story.append(table)
                story.append(Spacer(1, 8*mm))

            # --- Оборудование ---
            equipment = self.process_data.get('equipment', [])
            if equipment:
                data = [["Наименование", "Артикул", "Примечание"]]
                for _, name, art, note in equipment:
                    name_para = Paragraph(name, styles['CellText']) if len(name) > 30 else name
                    note_para = Paragraph(note, styles['CellText']) if note and len(note) > 30 else (note or "—")
                    data.append([name_para, art or "—", note_para])
                table = Table(data, colWidths=[70*mm, 50*mm, 60*mm], rowHeights=12*mm)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#9C27B0')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('FONTNAME', (0,0), (-1,0), font_name),
                    ('FONTSIZE', (0,0), (-1,0), 11),
                    ('FONTNAME', (0,1), (-1,-1), font_name),
                    ('FONTSIZE', (0,1), (-1,-1), 9),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                story.append(Paragraph("Оборудование", styles['Header']))
                story.append(table)

            # --- Подвал ---
            story.append(Spacer(1, 15*mm))
            story.append(Paragraph(f"Дата формирования: {self.process_data['timestamp']}", styles['Footer']))

            # --- Генерация с нумерацией ---
            doc.build(
                story,
                onFirstPage=self.add_page_number,
                onLaterPages=self.add_page_number
            )
            QMessageBox.information(self, "Успех", f"PDF сохранён:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось создать PDF:\n{e}")
            print(traceback.format_exc())

    def add_page_number(self, canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Страница {page_num}"
        canvas.setFont("DejaVu", 9)
        canvas.drawRightString(195*mm, 10*mm, text)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = CAPPWindow()
    window.show()
    sys.exit(app.exec_())
